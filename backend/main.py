from fastapi import FastAPI, File, UploadFile, HTTPException, Response
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import numpy as np
from datetime import datetime
import os
import shutil
import tempfile
import logging
import json
import traceback
import zipfile
from pathlib import Path
import sys
from typing import cast
import pandas as pd
from pandas import DataFrame, Series
from xml.etree import ElementTree as ET
import uuid

# Настройка логирования
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Логируем информацию о запуске
logger.info("=== Starting OFD Converter Backend ===")
logger.info(f"Python Version: {sys.version}")
logger.info(f"Current Directory: {os.getcwd()}")
logger.info(f"Directory Contents: {os.listdir('.')}")

# Логируем только важные переменные окружения
env_vars = {
    "VERCEL_ENV": os.getenv("VERCEL_ENV", "local"),
    "VERCEL_REGION": os.getenv("VERCEL_REGION", "unknown"),
    "PYTHON_VERSION": sys.version,
    "TEMP_DIR": tempfile.gettempdir()
}
logger.info(f"Environment Info: {json.dumps(env_vars, indent=2)}")

app = FastAPI()

# Настройка CORS
origins = [
    "http://localhost:3000",
    "https://ofd-converter.vercel.app",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

@app.get("/api")
async def root():
    logger.info("Root endpoint called")
    return {
        "status": "Backend is running",
        "timestamp": datetime.now().isoformat(),
        "environment": os.getenv("VERCEL_ENV", "local")
    }

@app.get("/api/health")
async def health_check():
    """Подробная проверка состояния бэкенда"""
    logger.info("Health check endpoint called")
    
    # Проверяем доступ к временной директории
    temp_dir = tempfile.gettempdir()
    temp_writable = os.access(temp_dir, os.W_OK)
    
    # Проверяем наличие всех необходимых пакетов
    required_packages = ['pandas', 'numpy', 'openpyxl']
    packages_status = {}
    for package in required_packages:
        try:
            module = __import__(package)
            packages_status[package] = getattr(module, '__version__', 'installed')
        except ImportError as e:
            packages_status[package] = f"ERROR: {str(e)}"
    
    # Собираем информацию о системе
    system_info = {
        "python_version": sys.version,
        "platform": sys.platform,
        "cwd": os.getcwd(),
        "temp_dir": temp_dir,
        "temp_dir_writable": temp_writable,
        "dir_contents": os.listdir('.'),
        "packages": packages_status,
        "environment": os.getenv("VERCEL_ENV", "local"),
        "vercel_region": os.getenv("VERCEL_REGION"),
        "memory_limit": os.getenv("AWS_LAMBDA_FUNCTION_MEMORY_SIZE"),
        "function_name": os.getenv("AWS_LAMBDA_FUNCTION_NAME"),
        "timestamp": datetime.now().isoformat()
    }
    
    logger.info(f"Health check response: {json.dumps(system_info, indent=2)}")
    return system_info

# Константы
PAYMENT_COLUMNS = ['Наличными', 'Электронными', 'Предоплата (аванс)', 'Зачет предоплаты (аванса)']
HIGHLIGHT_COLOR = 'D3D3D3'  # Светло-серый цвет для итоговых строк

# Определяем путь к временной директории
try:
    TEMP_DIR = "/tmp" if os.path.exists("/tmp") else "temp_files"
    if not os.path.exists(TEMP_DIR):
        os.makedirs(TEMP_DIR, exist_ok=True)
        logger.info(f"Created temporary directory: {TEMP_DIR}")
except Exception as e:
    logger.error(f"Failed to create temporary directory: {e}")
    raise HTTPException(
        status_code=500,
        detail="Failed to create temporary directory for file processing"
    )

def process_dataframe(df: DataFrame) -> DataFrame:
    """Обработка данных согласно требованиям"""
    # 3. Сортировка по дате
    df = df.copy()
    df['Дата/время'] = pd.to_datetime(df['Дата/время'])
    df = df.sort_values('Дата/время')
    return df

def add_daily_totals(df: DataFrame, writer: pd.ExcelWriter, sheet_name: str) -> None:
    """Добавление ежедневных итогов с форматированием"""
    logger.info(f"Adding daily totals for sheet: {sheet_name}")
    
    # Преобразуем столбец даты в datetime
    df['Дата/время'] = pd.to_datetime(df['Дата/время'])
    df['Дата'] = df['Дата/время'].dt.date
    
    # Группируем по дате и считаем итоги
    daily_totals = df.groupby('Дата').agg({
        col: 'sum' for col in PAYMENT_COLUMNS
    }).reset_index()
    
    # Записываем данные в Excel
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Получаем объект листа
    worksheet = writer.sheets[sheet_name]
    
    # Добавляем итоги после основных данных
    start_row = len(df) + 3
    daily_totals.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
    
    # Форматирование итогов
    for row in worksheet[start_row + 1:start_row + len(daily_totals) + 2]:
        for cell in row:
            cell.fill = PatternFill(start_color=HIGHLIGHT_COLOR, end_color=HIGHLIGHT_COLOR, fill_type='solid')

def process_nomenclature_dataframe(df: DataFrame) -> DataFrame:
    """Обработка данных для отчета по номенклатуре"""
    logger.info("Processing nomenclature report")
    
    # Создаем копию DataFrame
    df = df.copy()
    
    # Обработка предоплаты
    prepayment_column = 'Зачет предоплаты (аванса) по чеку'
    if prepayment_column in df.columns:
        # Заполняем NaN значения нулями
        df[prepayment_column] = df[prepayment_column].fillna(0)
        
        # Группируем по номеру чека для обработки предоплаты
        receipt_groups = df.groupby('Номер документа')
        
        for receipt_num, receipt_df in receipt_groups:
            # Если есть предоплата в чеке
            if (receipt_df[prepayment_column] > 0).any():
                # Получаем общую сумму предоплаты для чека (берем только одно значение)
                remaining_prepayment = receipt_df[prepayment_column].iloc[0]
                
                # Создаем маску для текущего чека
                receipt_mask = df['Номер документа'] == receipt_num
                
                # Обрабатываем каждую позицию в чеке
                for idx in receipt_df.index:
                    current_amount = df.loc[idx, 'Сумма товара']
                    
                    if remaining_prepayment <= 0:
                        break
                        
                    if current_amount >= remaining_prepayment:
                        # Случай 1: Сумма товара больше или равна остатку предоплаты
                        df.loc[idx, 'Сумма товара'] = current_amount - remaining_prepayment
                        # Обнуляем остальные одинаковые позиции в чеке
                        same_items_mask = (df['Номер документа'] == receipt_num) & \
                                        (df.index > idx) & \
                                        (df['Наименование'] == df.loc[idx, 'Наименование'])
                        df.loc[same_items_mask, 'Сумма товара'] = 0
                        remaining_prepayment = 0
                    else:
                        # Случай 2: Предоплата больше суммы товара
                        df.loc[idx, 'Сумма товара'] = 0
                        remaining_prepayment -= current_amount
                        
    # Обработка значений согласно правилам
    for column in ['Наличными по чеку', 'Электронными по чеку']:
        # Замена значений, которые больше 'Сумма товара'
        mask = (df[column] > df['Сумма товара']) & (df[column] > 0)
        df.loc[mask, column] = df.loc[mask, 'Сумма товара']
    
    # Обработка возвратов
    mask_return = df['Признак расчета (тег 1054)'] == 'Возврат прихода'
    for column in ['Наличными по чеку', 'Электронными по чеку', 'Сумма товара']:
        df.loc[mask_return, column] = -df.loc[mask_return, column]
    
    # После всех расчетов устанавливаем 'Сумма товара' равной сумме наличных и электронных платежей
    df['Сумма товара'] = df['Наличными по чеку'] + df['Электронными по чеку']
    
    return df

def add_daily_totals_nomenclature(df: DataFrame, writer: pd.ExcelWriter, sheet_name: str) -> None:
    """Добавление ежедневных итогов для отчета по номенклатуре"""
    logger.info(f"Adding daily totals for sheet: {sheet_name}")
    
    # Преобразуем столбец даты в datetime
    df['Дата/время'] = pd.to_datetime(df['Дата/время'])
    df['Дата'] = df['Дата/время'].dt.date
    
    # Группируем по дате и считаем итоги
    daily_totals = df.groupby('Дата').agg({
        'Наличными по чеку': 'sum',
        'Электронными по чеку': 'sum',
        'Сумма товара': 'sum'
    }).reset_index()
    
    # Записываем данные в Excel
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Получаем объект листа
    worksheet = writer.sheets[sheet_name]
    
    # Добавляем итоги после основных данных
    start_row = len(df) + 3
    daily_totals.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
    
    # Форматирование итогов
    fill = PatternFill(start_color=HIGHLIGHT_COLOR, end_color=HIGHLIGHT_COLOR, fill_type='solid')
    for row in range(start_row + 1, start_row + len(daily_totals) + 2):
        for col in range(1, len(daily_totals.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill

def process_taxcom_dataframe(df: DataFrame) -> DataFrame:
    """Обработка данных для Такском отчета по чекам"""
    logger.info("Processing taxcom report")
    
    # Создаем копию DataFrame
    df = df.copy()
    
    # Удаляем итоговые строки
    df = df[~df['Дата и время'].astype(str).str.contains('Итог', case=False, na=False)]
    
    # Преобразуем и сортируем по дате
    df['Дата и время'] = pd.to_datetime(df['Дата и время'])
    df = df.sort_values('Дата и время')
    
    return df

def add_daily_totals_taxcom(df: DataFrame, writer: pd.ExcelWriter, sheet_name: str) -> None:
    """Добавление ежедневных итогов для Такском отчета"""
    logger.info(f"Adding daily totals for taxcom sheet: {sheet_name}")
    
    # Создаем копию DataFrame для работы
    df = df.copy()
    
    # Преобразуем столбец даты в datetime и создаем столбец только с датой
    df['Дата'] = df['Дата и время'].dt.date
    
    # Группируем по дате и считаем итоги
    daily_totals = df.groupby('Дата').agg({
        'Наличными': 'sum',
        'Безналичными': 'sum',
        'Сумма': 'sum'
    }).reset_index()
    
    # Форматируем даты в строки для вывода
    daily_totals['Дата'] = daily_totals['Дата'].astype(str)
    
    # Добавляем строку с общим итогом
    total_row = pd.DataFrame([{
        'Дата': 'Итог',
        'Наличными': daily_totals['Наличными'].sum(),
        'Безналичными': daily_totals['Безналичными'].sum(),
        'Сумма': daily_totals['Сумма'].sum()
    }])
    daily_totals = pd.concat([daily_totals, total_row], ignore_index=True)
    
    # Записываем основные данные
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Получаем объект листа
    worksheet = writer.sheets[sheet_name]
    
    # Добавляем итоги после основных данных
    start_row = len(df) + 3
    daily_totals.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
    
    # Форматирование итогов
    fill = PatternFill(start_color=HIGHLIGHT_COLOR, end_color=HIGHLIGHT_COLOR, fill_type='solid')
    for row in range(start_row + 1, start_row + len(daily_totals) + 2):
        for col in range(1, len(daily_totals.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill

def create_card_xml(source_xml: ET.Element) -> ET.Element:
    """Создает card.xml на основе данных из исходного файла"""
    # Создаем корневой элемент Card
    card = ET.Element("Card", {
        "xmlns": "http://api-invoice.taxcom.ru/card",
        "xmlns:xs": "http://www.w3.org/2001/XMLSchema",
        "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
        "DocumentVersion": "1.0"
    })

    # Добавляем Identifiers с уникальным ExternalIdentifier
    identifiers = ET.SubElement(card, "Identifiers")
    external_id = str(uuid.uuid4())
    identifiers.set("ExternalIdentifier", external_id)

    # Добавляем Type с информацией о типе документа
    type_elem = ET.SubElement(card, "Type")
    type_elem.text = "Invoice"  # Изменили тип на Invoice для счета на оплату

    # Добавляем Description
    description = ET.SubElement(card, "Description")
    title = "Счет на оплату"  # Значение по умолчанию
    date = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")  # Значение по умолчанию
    doc_number = ""  # Значение по умолчанию для номера документа

    try:
        # Получаем данные из исходного XML
        doc = source_xml.find(".//Документ")
        if doc is not None:
            title = doc.get("НаимДокОпр", title)
            doc_number = doc.get("НомерСчФ") or doc.get("НомИнфПр", "")
            date_str = doc.get("ДатаИнфПр") or doc.get("ДатаСчФ")
            if date_str:
                # Преобразуем дату в нужный формат
                date_obj = datetime.strptime(date_str, "%d.%m.%Y")
                date = date_obj.strftime("%Y-%m-%dT%H:%M:%S")
    except (AttributeError, ValueError) as e:
        logger.warning(f"Ошибка при получении данных для Description: {e}")

    description.set("Title", title)
    description.set("Date", date)
    if doc_number:
        description.set("Number", doc_number)

    # Добавляем Direction
    direction = ET.SubElement(card, "Direction")
    direction.text = "Outbound"  # Исходящий документ

    # Добавляем Sender
    sender = ET.SubElement(card, "Sender")
    try:
        # Получаем данные о продавце из исходного XML
        seller = source_xml.find(".//СвПрод")
        if seller is not None:
            abonent = ET.SubElement(sender, "Abonent")
            inn = seller.get("ИННЮЛ", "") or seller.get("ИННФЛ", "")
            kpp = seller.get("КПП", "")
            name = seller.get("НаимОрг", "") or f"ИП {seller.get('ФИО', '')}"
            
            abonent.set("Id", inn)  # Используем ИНН как идентификатор
            abonent.set("Name", name)
            abonent.set("Inn", inn)
            if kpp:
                abonent.set("Kpp", kpp)
            
            # Добавляем Department для Sender
            department = ET.SubElement(sender, "Department")
            department.set("Id", "0")  # Значение по умолчанию для основного подразделения
    except Exception as e:
        logger.warning(f"Ошибка при получении данных отправителя: {e}")

    # Добавляем Receiver
    receiver = ET.SubElement(card, "Receiver")
    try:
        # Получаем данные о покупателе из исходного XML
        buyer = source_xml.find(".//СвПокуп")
        if buyer is not None:
            abonent = ET.SubElement(receiver, "Abonent")
            inn = buyer.get("ИННЮЛ", "") or buyer.get("ИННФЛ", "")
            kpp = buyer.get("КПП", "")
            name = buyer.get("НаимОрг", "") or f"ИП {buyer.get('ФИО', '')}"
            
            abonent.set("Id", inn)  # Используем ИНН как идентификатор
            abonent.set("Name", name)
            abonent.set("Inn", inn)
            if kpp:
                abonent.set("Kpp", kpp)
            
            # Добавляем Department для Receiver
            department = ET.SubElement(receiver, "Department")
            department.set("Id", "0")  # Значение по умолчанию для основного подразделения
    except Exception as e:
        logger.warning(f"Ошибка при получении данных получателя: {e}")

    # Добавляем DocumentState
    state = ET.SubElement(card, "DocumentState")
    state.text = "Sent"  # Статус документа

    return card

def create_meta_xml(source_xml: ET.Element) -> ET.Element:
    """Создает meta.xml на основе данных из исходного файла"""
    # Создаем корневой элемент ContainerDescription
    container = ET.Element("ContainerDescription", {
        "xmlns": "http://api-invoice.taxcom.ru/meta",
        "xmlns:xs": "http://www.w3.org/2001/XMLSchema",
        "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
        "DocumentVersion": "1.0"
    })

    # Добавляем DocFlow с уникальным Id
    doc_flow = ET.SubElement(container, "DocFlow")
    doc_flow.set("Id", str(uuid.uuid4()))
    doc_flow.set("DocumentCount", "1")

    # Добавляем Documents
    documents = ET.SubElement(doc_flow, "Documents")
    document = ET.SubElement(documents, "Document")
    
    # Устанавливаем атрибуты документа
    try:
        doc_type = source_xml.find(".//Документ").get("Функция", "")
        if doc_type == "СЧФ":
            document.set("ReglamentCode", "Invoice")
        else:
            document.set("ReglamentCode", "Nonformalized")
    except (AttributeError, ValueError):
        document.set("ReglamentCode", "Invoice")  # По умолчанию для счета на оплату
    
    document.set("TransactionCode", "MainDocument")
    document.set("DocumentDate", datetime.now().strftime("%Y-%m-%dT%H:%M:%S"))
    document.set("DocumentNumber", "1")

    # Добавляем Files
    files = ET.SubElement(document, "Files")

    # Добавляем MainImage
    main_image = ET.SubElement(files, "MainImage")
    main_image.set("xmlns:d6p1", "http://api-invoice.taxcom.ru/card")
    
    # Получаем имя исходного файла
    try:
        source_filename = source_xml.find(".//Файл").get("ИмяФайл", "document.xml")
    except (AttributeError, ValueError):
        source_filename = "document.xml"
    main_image.set("Path", f"1/{source_filename}")

    # Добавляем ExternalCard
    external_card = ET.SubElement(files, "ExternalCard")
    external_card.set("xmlns:d6p1", "http://api-invoice.taxcom.ru/card")
    external_card.set("Path", "1/card.xml")

    # Добавляем ProcessingState
    state = ET.SubElement(document, "ProcessingState")
    state.text = "New"

    return container

@app.post("/api/process_excel")
async def process_excel(file: UploadFile = File(...), report_type: str = 'checks'):
    temp_path = None
    output_files = []
    archive_name = None
    
    try:
        logger.info(f"Получен файл: {file.filename}, тип отчета: {report_type}")
        
        # Проверка наличия файла и его имени
        if not file or not file.filename:
            raise HTTPException(status_code=400, detail="No file provided")
            
        # Проверка расширения файла
        if not str(file.filename).endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

        # Генерируем уникальные имена файлов
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_path = os.path.join(TEMP_DIR, f"temp_{timestamp}_{file.filename}")
        
        # Сохраняем входной файл
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        logger.info("File saved successfully")

        # Читаем Excel файл
        logger.info("Reading Excel file")
        df = cast(DataFrame, pd.read_excel(temp_path))
        logger.info(f"DataFrame shape: {df.shape}")

        # Определяем тип отчета на основе наличия колонок
        logger.info(f"Detecting report type based on columns")
        checks_columns = ['Признак расчета', 'Тип налогообложения']
        nomenclature_columns = ['Признак расчета (тег 1054)', 'Признак предмета расчета (тег 1212)']
        taxcom_columns = ['Дата и время', 'Система налогообложения', 'Наличными', 'Безналичными', 'Сумма']

        has_checks_columns = all(col in df.columns for col in checks_columns)
        has_nomenclature_columns = all(col in df.columns for col in nomenclature_columns)
        has_taxcom_columns = all(col in df.columns for col in taxcom_columns)

        # Автоматически определяем тип отчета, если он не соответствует структуре
        detected_type = report_type
        if report_type == 'checks' and not has_checks_columns:
            if has_nomenclature_columns:
                detected_type = 'nomenclature'
            elif has_taxcom_columns:
                detected_type = 'taxcom'
        elif report_type == 'nomenclature' and not has_nomenclature_columns:
            if has_checks_columns:
                detected_type = 'checks'
            elif has_taxcom_columns:
                detected_type = 'taxcom'
        elif report_type == 'taxcom' and not has_taxcom_columns:
            if has_checks_columns:
                detected_type = 'checks'
            elif has_nomenclature_columns:
                detected_type = 'nomenclature'

        # Проверяем наличие необходимых колонок в зависимости от типа отчета
        if detected_type == 'checks':
            required_columns = ['Дата/время', 'Признак расчета', 'Тип налогообложения']
        elif detected_type == 'nomenclature':
            required_columns = [
                'Дата/время', 'Признак расчета (тег 1054)', 'Признак предмета расчета (тег 1212)',
                'Наличными по чеку', 'Электронными по чеку', 'Сумма товара'
            ]
        else:  # taxcom
            required_columns = ['Дата и время', 'Система налогообложения', 'Наличными', 'Безналичными', 'Сумма']
            
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns for {detected_type} report: {', '.join(missing_columns)}"
            )

        # Обрабатываем данные в зависимости от типа отчета
        logger.info(f"Processing data for report type: {detected_type}")
        if detected_type == 'checks':
            df = process_dataframe(df)
            # Разделяем по типу налогообложения
            for tax_type in ['ПАТЕНТ', 'УСН']:
                mask = df['Тип налогообложения'].str.contains(tax_type, case=False, na=False)
                df_filtered = cast(DataFrame, df[mask])
                if not df_filtered.empty:
                    output_filename = os.path.join(TEMP_DIR, f"processed_{tax_type}_{timestamp}_{file.filename}")
                    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                        add_daily_totals(df_filtered.copy(), writer, f'{tax_type}')
                    output_files.append(output_filename)
        elif detected_type == 'nomenclature':
            df = process_nomenclature_dataframe(df)
            # Разделяем по признаку предмета расчета
            for item_type in df['Признак предмета расчета (тег 1212)'].unique():
                if pd.isna(item_type):
                    continue
                mask = df['Признак предмета расчета (тег 1212)'] == item_type
                df_filtered = cast(DataFrame, df[mask])
                if not df_filtered.empty:
                    safe_item_type = "".join(x for x in str(item_type) if x.isalnum() or x in (' ', '-', '_'))[:50]
                    output_filename = os.path.join(TEMP_DIR, f"processed_{safe_item_type}_{timestamp}_{file.filename}")
                    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                        add_daily_totals_nomenclature(df_filtered.copy(), writer, safe_item_type)
                    output_files.append(output_filename)
        else:  # taxcom
            df = process_taxcom_dataframe(df)
            # Разделяем по системе налогообложения
            tax_types_map = {'Патент': 'PATENT', 'УСН доход': 'USN'}
            for tax_type, file_suffix in tax_types_map.items():
                mask = df['Система налогообложения'] == tax_type
                df_filtered = cast(DataFrame, df[mask])
                if not df_filtered.empty:
                    output_filename = os.path.join(TEMP_DIR, f"processed_{file_suffix}_{timestamp}_{file.filename}")
                    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                        add_daily_totals_taxcom(df_filtered.copy(), writer, tax_type)
                    output_files.append(output_filename)
        
        # Проверяем, что файлы созданы
        if not output_files:
            raise Exception("Не удалось создать выходные файлы")
        
        # Создаем архив с результатами
        archive_name = os.path.join(TEMP_DIR, f"results_{timestamp}.zip")
        logger.info(f"Creating ZIP archive: {archive_name}")
        
        with zipfile.ZipFile(archive_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for f in output_files:
                if os.path.exists(f):
                    zipf.write(f, os.path.basename(f))
                else:
                    logger.error(f"Файл {f} не найден при создании архива")
        
        # Проверяем, что архив создан
        if not os.path.exists(archive_name):
            raise Exception("Не удалось создать архив с результатами")
        
        logger.info("Processing completed successfully")
        
        # Читаем архив в память
        with open(archive_name, 'rb') as f:
            file_data = f.read()
        
        # Удаляем временные файлы
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)
        for f in output_files:
            if os.path.exists(f):
                os.remove(f)
        if archive_name and os.path.exists(archive_name):
            os.remove(archive_name)
        
        # Возвращаем архив с правильными заголовками
        return Response(
            content=file_data,
            media_type='application/zip',
            headers={
                'Content-Disposition': f'attachment; filename="results_{timestamp}.zip"',
                'Content-Type': 'application/zip'
            }
        )
    
    except Exception as e:
        logger.error(f"Error processing file: {str(e)}", exc_info=True)
        # В случае ошибки очищаем все файлы
        try:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)
            for f in output_files:
                if os.path.exists(f):
                    os.remove(f)
            if archive_name and os.path.exists(archive_name):
                os.remove(archive_name)
        except Exception as cleanup_error:
            logger.error(f"Error cleaning up files: {str(cleanup_error)}")
        
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/process_bill")
async def process_bill(file: UploadFile = File(...)):
    """Обработка электронного счета"""
    temp_dir = None
    archive_name = None
    response = None
    
    try:
        logger.info(f"Processing electronic bill: {file.filename}")
        
        # Проверка расширения файла
        if not file.filename.lower().endswith('.xml'):
            raise HTTPException(status_code=400, detail="Only XML files are allowed")
        
        # Создаем временную директорию для работы с файлами
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_dir = os.path.join(TEMP_DIR, f"bill_processing_{timestamp}")
        os.makedirs(temp_dir)
        logger.info(f"Created temp directory: {temp_dir}")
        
        # Создаем структуру папок
        bill_dir = os.path.join(temp_dir, "1")
        os.makedirs(bill_dir)
        logger.info(f"Created bill directory: {bill_dir}")
        
        # Читаем входной XML файл
        content = await file.read()
        logger.info(f"Read file content, size: {len(content)} bytes")
        
        # Определяем кодировку файла
        encoding = 'utf-8'
        if content.startswith(b'\xef\xbb\xbf'):  # UTF-8 с BOM
            content = content[3:]
            logger.info("Detected UTF-8 with BOM")
        elif b'windows-1251' in content.lower() or b'cp1251' in content.lower():
            encoding = 'windows-1251'
            logger.info("Detected windows-1251 encoding")
        
        try:
            # Пробуем декодировать XML с определенной кодировкой
            xml_content = content.decode(encoding)
            logger.info(f"Successfully decoded content with {encoding}")
            
            # Логируем первые 200 символов содержимого для отладки
            logger.info(f"Content preview: {xml_content[:200]}")
            
            source_xml = ET.fromstring(xml_content)
            logger.info("Successfully parsed XML")
            
        except (UnicodeDecodeError, ET.ParseError) as e:
            logger.warning(f"Failed to decode with {encoding}: {str(e)}")
            # Если не удалось, пробуем другие кодировки
            encodings = ['windows-1251', 'utf-8', 'utf-16', 'cp866']
            for enc in encodings:
                if enc != encoding:
                    try:
                        xml_content = content.decode(enc)
                        source_xml = ET.fromstring(xml_content)
                        encoding = enc
                        logger.info(f"Successfully decoded with alternative encoding: {enc}")
                        break
                    except (UnicodeDecodeError, ET.ParseError) as e:
                        logger.warning(f"Failed to decode with {enc}: {str(e)}")
                        continue
            else:
                logger.error("Failed to decode with any encoding")
                raise HTTPException(
                    status_code=400,
                    detail="Не удалось определить кодировку файла или файл содержит некорректный XML"
                )
        
        try:
            # Сохраняем исходный файл
            source_path = os.path.join(bill_dir, file.filename)
            with open(source_path, 'w', encoding='windows-1251') as f:
                f.write(xml_content)
            logger.info(f"Saved source file: {source_path}")
            
            # Создаем card.xml
            logger.info("Creating card.xml")
            card_xml = create_card_xml(source_xml)
            card_content = ('<?xml version="1.0" encoding="windows-1251"?>\n' + 
                          ET.tostring(card_xml, encoding='unicode'))
            card_path = os.path.join(bill_dir, 'card.xml')
            with open(card_path, 'w', encoding='windows-1251') as f:
                f.write(card_content)
            logger.info(f"Saved card.xml: {card_path}")
            
            # Создаем meta.xml
            logger.info("Creating meta.xml")
            meta_xml = create_meta_xml(source_xml)
            meta_content = ('<?xml version="1.0" encoding="windows-1251"?>\n' + 
                          ET.tostring(meta_xml, encoding='unicode'))
            meta_path = os.path.join(temp_dir, 'meta.xml')
            with open(meta_path, 'w', encoding='windows-1251') as f:
                f.write(meta_content)
            logger.info(f"Saved meta.xml: {meta_path}")
            
            # Создаем ZIP архив
            logger.info("Creating ZIP archive")
            archive_name = os.path.join(TEMP_DIR, f"bill_{timestamp}.zip")
            with zipfile.ZipFile(archive_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Добавляем meta.xml в корень архива
                zipf.write(meta_path, 'meta.xml')
                # Добавляем файлы из папки 1
                for root, _, files in os.walk(bill_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.join('1', file)
                        zipf.write(file_path, arcname)
            logger.info(f"Created ZIP archive: {archive_name}")
            
            # Проверяем, что архив существует и имеет размер
            if not os.path.exists(archive_name):
                raise HTTPException(
                    status_code=500,
                    detail="Ошибка при создании архива: файл не найден"
                )
                
            archive_size = os.path.getsize(archive_name)
            logger.info(f"Archive size: {archive_size} bytes")
            
            if archive_size == 0:
                raise HTTPException(
                    status_code=500,
                    detail="Ошибка при создании архива: файл пуст"
                )
            
            # Читаем архив в память перед отправкой
            with open(archive_name, 'rb') as f:
                archive_data = f.read()
                
            # Создаем StreamingResponse
            response = StreamingResponse(
                iter([archive_data]),
                media_type='application/zip',
                headers={
                    'Content-Disposition': f'attachment; filename="bill_{timestamp}.zip"',
                    'Content-Length': str(len(archive_data))
                }
            )
            
            # Очищаем временные файлы
            if temp_dir and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                logger.info(f"Cleaned up temp directory: {temp_dir}")
            
            return response
            
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Unhandled error: {str(e)}", exc_info=True)
            raise
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unhandled error: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка при обработке файла: {str(e)}"
        )
        
    finally:
        # В finally очищаем только архив, так как он уже прочитан в память
        try:
            if archive_name and os.path.exists(archive_name):
                os.remove(archive_name)
                logger.info(f"Cleaned up archive: {archive_name}")
        except Exception as e:
            logger.error(f"Error cleaning up archive: {str(e)}", exc_info=True)

@app.on_event("shutdown")
async def cleanup_temp_files():
    """Очистка временных файлов при выключении сервера"""
    if os.path.exists(TEMP_DIR):
        try:
            shutil.rmtree(TEMP_DIR)
            logger.info("Временная директория очищена")
        except Exception as e:
            logger.error(f"Ошибка при очистке временной директории: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    logger.info("Starting server")
    uvicorn.run(app, host="0.0.0.0", port=8000)
